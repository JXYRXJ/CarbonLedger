import io
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
import pytest
from app.services.export import ExportService


def test_csv_export():
    exporter = ExportService()
    headers = ["ColA", "ColB"]
    rows = [["Val1", 10.5], ["Val2", None]]
    
    csv_bytes = exporter.export_to_csv(headers, rows)
    assert isinstance(csv_bytes, bytes)
    
    csv_str = csv_bytes.decode("utf-8")
    assert "ColA,ColB" in csv_str
    assert "Val1,10.5" in csv_str
    assert "Val2" in csv_str


def test_excel_export():
    exporter = ExportService()
    headers = ["Name", "Amount"]
    rows = [["Alice", 100.50], ["Bob", 200.75]]
    
    excel_bytes = exporter.export_to_excel("TestSheet", headers, rows, totals_cols=[1])
    assert isinstance(excel_bytes, bytes)
    
    # Load workbook using openpyxl from memory buffer to check sheet and cells
    wb = load_workbook(io.BytesIO(excel_bytes))
    assert "TestSheet" in wb.sheetnames
    
    sheet = wb["TestSheet"]
    assert sheet.cell(row=1, column=1).value == "Name"
    assert sheet.cell(row=2, column=1).value == "Alice"
    assert sheet.cell(row=2, column=2).value == 100.50
    
    # Check totals row formula
    assert sheet.cell(row=4, column=1).value == "Total"
    assert sheet.cell(row=4, column=2).value == "=SUM(B2:B3)"


def test_pdf_export():
    exporter = ExportService()
    headers = ["Header1", "Header2"]
    rows = [["A1", "B1"], ["A2", "B2"]]
    
    pdf_bytes = exporter.export_to_pdf(
        title="Test PDF Title",
        headers=headers,
        rows=rows,
        company_info={"name": "Acme Corp", "id": "123-456"}
    )
    assert isinstance(pdf_bytes, bytes)
    # PDF files start with %PDF
    assert pdf_bytes.startswith(b"%PDF")
