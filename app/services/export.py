import io
import csv
from typing import List, Dict, Any, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


class ExportService:
    """
    Service generating multi-format reports (CSV, styled Excel workbooks via openpyxl, and printable PDF layouts via ReportLab).
    """

    def export_to_csv(self, headers: List[str], rows: List[List[Any]]) -> bytes:
        """
        Generates UTF-8 encoded CSV.
        """
        output = io.StringIO()
        writer = csv.writer(output, csv.excel)
        writer.writerow(headers)
        for row in rows:
            # Normalize dates/decimals for serialization
            writer.writerow([str(x) if x is not None else "" for x in row])
        return output.getvalue().encode("utf-8")

    def export_to_excel(
        self,
        sheet_name: str,
        headers: List[str],
        rows: List[List[Any]],
        totals_cols: Optional[List[int]] = None
    ) -> bytes:
        """
        Generates professionally designed Excel spreadsheet.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:30]  # Excel sheet title length limit is 31
        
        # Color Palettes (Sleek Charcoal & Navy design)
        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
        total_fill = PatternFill(start_color="E6ECF5", end_color="E6ECF5", fill_type="solid")
        
        thin_side = Side(border_style="thin", color="D1D5DB")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        double_bottom = Border(top=thin_side, bottom=Side(border_style="double", color="1B365D"))
        
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = cell_border
        
        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            is_even = row_idx % 2 == 0
            formatted_row = []
            for item in row_data:
                if isinstance(item, (int, float)):
                    formatted_row.append(item)
                elif isinstance(item, datetime):
                    formatted_row.append(item.strftime("%Y-%m-%d %H:%M"))
                else:
                    formatted_row.append(str(item) if item is not None else "")
            
            ws.append(formatted_row)
            
            # Format styles
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = cell_border
                if is_even:
                    cell.fill = zebra_fill
                
                # Check data types for alignments/number format
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.alignment = align_right
                    cell.number_format = "#,##0.00"
                else:
                    cell.alignment = align_left
        
        # Add totals row if specified
        if totals_cols and rows:
            tot_row_idx = len(rows) + 2
            tot_row = ["Total" if i == 0 else "" for i in range(len(headers))]
            ws.append(tot_row)
            
            # Style Totals row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=tot_row_idx, column=col_idx)
                cell.fill = total_fill
                cell.font = Font(name="Calibri", size=11, bold=True)
                cell.border = double_bottom
                
                if col_idx - 1 in totals_cols:
                    col_letter = get_column_letter(col_idx)
                    cell.value = f"=SUM({col_letter}2:{col_letter}{tot_row_idx - 1})"
                    cell.number_format = "#,##0.00"
                    cell.alignment = align_right
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or '')
                if cell.value and val_str.startswith('='):
                    val_str = "Total: 1,000,000.00"  # mock calculation length
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        # Save workbook to memory
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_to_pdf(
        self,
        title: str,
        headers: List[str],
        rows: List[List[Any]],
        company_info: Optional[Dict[str, str]] = None
    ) -> bytes:
        """
        Generates formal Report PDF.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            name="ReportTitle",
            parent=styles["Heading1"],
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#1B365D"),
            spaceAfter=15
        )
        meta_style = ParagraphStyle(
            name="ReportMeta",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#4B5563"),
            spaceAfter=15
        )
        table_cell_style = ParagraphStyle(
            name="TableCell",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
            textColor=colors.HexColor("#1F2937")
        )
        table_header_style = ParagraphStyle(
            name="TableHeader",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
            bold=True,
            textColor=colors.white
        )
        
        story = []
        
        # Add Title
        story.append(Paragraph(title, title_style))
        
        # Add Metadata
        meta_text = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
        if company_info:
            meta_text += f"Company: {company_info.get('name', 'N/A')} (ID: {company_info.get('id', 'N/A')})<br/>"
        story.append(Paragraph(meta_text, meta_style))
        story.append(Spacer(1, 10))
        
        # Prep Table Data
        pdf_table_data = []
        
        # Header Row
        pdf_table_data.append([Paragraph(h, table_header_style) for h in headers])
        
        # Data Rows
        for row in rows:
            pdf_row = []
            for item in row:
                if isinstance(item, (int, float)):
                    item_str = f"{item:,.2f}"
                elif isinstance(item, datetime):
                    item_str = item.strftime("%Y-%m-%d %H:%M")
                else:
                    item_str = str(item) if item is not None else ""
                pdf_row.append(Paragraph(item_str, table_cell_style))
            pdf_table_data.append(pdf_row)
            
        # Determine column widths
        col_width = (doc.width) / len(headers)
        
        t = Table(pdf_table_data, colWidths=[col_width] * len(headers))
        
        # Premium Table styling (alternating row backgrounds)
        t_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B365D")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ])
        
        # Alternate background colors
        for i in range(1, len(pdf_table_data)):
            if i % 2 == 0:
                t_style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F9FAFB"))
                
        t.setStyle(t_style)
        story.append(t)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
